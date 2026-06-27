from rest_framework import generics, filters
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
import openpyxl
from django.http import HttpResponse

from .models import Youth
from .serializers import YouthSerializer
from apps.accounts.models import CustomUser, Organization, District


def _notify_yetakchi_telegram(meeting, youth, rahbar):
    """Yetakchiga Telegram xabar yuborish — uchrashuv bo'lganda."""
    import requests, os
    token = os.environ.get('BOT_TOKEN', '')
    yetakchi = youth.yetakchi
    if not token or not yetakchi or not yetakchi.telegram_id:
        return
    text = (
        f"🤝 <b>Yangi uchrashuv!</b>\n\n"
        f"👤 Yosh: <b>{youth.full_name}</b>\n"
        f"👨‍💼 Rahbar: {rahbar.get_full_name()}\n"
        f"📅 Sana: {meeting.date.strftime('%d.%m.%Y %H:%M')}"
    )
    result_url = f"https://sam-auth.uz/uchrashuvlar/{meeting.id}/natija/"
    keyboard = {"inline_keyboard": [[{"text": "📋 Natijani ko'rish", "url": result_url}]]}
    try:
        requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": yetakchi.telegram_id, "text": text,
                  "parse_mode": "HTML", "reply_markup": keyboard},
            timeout=5
        )
    except Exception:
        pass


class YouthListAPI(generics.ListAPIView):
    serializer_class = YouthSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['full_name', 'address', 'phone']

    def get_queryset(self):
        user = self.request.user
        qs = Youth.objects.filter(is_active=True)
        if user.role == 'rahbar':
            return qs.filter(rahbar=user)
        if user.role == 'yetakchi':
            return qs.filter(yetakchi=user)
        return qs


class YouthDetailAPI(generics.RetrieveAPIView):
    serializer_class = YouthSerializer

    def get_queryset(self):
        user = self.request.user
        qs = Youth.objects.all()
        if user.role == 'rahbar':
            return qs.filter(rahbar=user)
        if user.role == 'yetakchi':
            return qs.filter(yetakchi=user)
        return qs


@login_required
def youth_list(request):
    user = request.user
    qs = Youth.objects.filter(is_active=True).select_related('organization', 'organization__district', 'rahbar', 'yetakchi')
    if user.role == 'rahbar':
        qs = qs.filter(rahbar=user)
    elif user.role == 'yetakchi':
        qs = qs.filter(yetakchi=user)

    search = request.GET.get('q', '')
    tizim_filter = request.GET.get('tizim', '')
    district_filter = request.GET.get('district', '')

    if search:
        qs = qs.filter(full_name__icontains=search)
    if tizim_filter:
        qs = qs.filter(notes__icontains=f'Biriktirilgan tizim: {tizim_filter}')
    if district_filter:
        qs = qs.filter(organization__district_id=district_filter)

    total = qs.count()
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page', 1))

    districts = District.objects.all() if user.is_admin else District.objects.none()

    # Unique tizim qiymatlarini notes dan ajratib olish
    tizim_values = set()
    for notes in Youth.objects.filter(notes__icontains='Biriktirilgan tizim:').values_list('notes', flat=True):
        for line in (notes or '').split('\n'):
            line = line.strip()
            if line.startswith('Biriktirilgan tizim:'):
                val = line.replace('Biriktirilgan tizim:', '').strip()
                if val:
                    tizim_values.add(val)
    tizim_values = sorted(tizim_values)

    from apps.surveys.models import Survey
    has_active_survey = Survey.objects.filter(is_active=True).exists()

    return render(request, 'youth/list.html', {
        'page': page,
        'search': search,
        'total': total,
        'tizim_filter': tizim_filter,
        'district_filter': district_filter,
        'tizim_values': tizim_values,
        'districts': districts,
        'has_active_survey': has_active_survey,
    })


@login_required
def youth_detail(request, pk):
    from apps.surveys.models import Survey
    youth = get_object_or_404(Youth, pk=pk)
    meetings = youth.meetings.order_by('-created_at')[:10]
    has_active_survey = Survey.objects.filter(is_active=True).exists()
    return render(request, 'youth/detail.html', {
        'youth': youth,
        'meetings': meetings,
        'has_active_survey': has_active_survey,
    })


@login_required
def youth_survey_view(request, pk):
    """Web so'rovnoma formasi — rahbar uchun."""
    from apps.surveys.models import Survey
    from apps.meetings.models import Meeting, Verification
    from django.utils import timezone
    import json

    youth = get_object_or_404(Youth, pk=pk)

    if request.user.role == 'rahbar' and youth.rahbar_id != request.user.id:
        return HttpResponse("Ruxsat yo'q", status=403)
    if request.user.role == 'yetakchi' and youth.yetakchi_id != request.user.id:
        return HttpResponse("Ruxsat yo'q", status=403)

    survey = Survey.objects.filter(is_active=True).prefetch_related('questions').first()
    if not survey:
        return render(request, 'meetings/survey_no_active.html', {'youth': youth})

    questions = list(survey.questions.order_by('order'))

    if request.method == 'POST':
        answers = []
        lat, lng = None, None
        photo_file = None
        notes_parts = []

        for q in questions:
            qtype = q.type
            if qtype == 'location':
                lat_v = request.POST.get(f'q_{q.id}_lat', '').strip()
                lng_v = request.POST.get(f'q_{q.id}_lng', '').strip()
                if lat_v and lng_v:
                    try:
                        lat, lng = float(lat_v), float(lng_v)
                        val = f"{lat:.6f}, {lng:.6f}"
                        notes_parts.append(f"📍 {q.text}: {val}")
                        answers.append({'q': q.text, 'type': 'location', 'value': val})
                    except ValueError:
                        pass
                elif q.required:
                    pass  # Majburiy emas bo'lsa o'tkazib yuborish
            elif qtype == 'photo':
                pf = request.FILES.get(f'q_{q.id}_photo')
                if pf:
                    photo_file = pf
                    notes_parts.append(f"📸 {q.text}: [Rasm]")
                    answers.append({'q': q.text, 'type': 'photo', 'value': 'rasm_yuklandi'})
            else:
                val = request.POST.get(f'q_{q.id}', '').strip()
                if val:
                    icon = '🔢' if qtype == 'number' else '✅' if qtype == 'choice' else '📝'
                    notes_parts.append(f"{icon} {q.text}: {val}")
                    answers.append({'q': q.text, 'type': qtype, 'value': val})

        answers_json = json.dumps(answers, ensure_ascii=False)
        notes_text = '\n'.join(notes_parts) + '\n[ANSWERS_JSON]\n' + answers_json

        # Rahbar aniqlash
        rahbar = request.user if request.user.role == 'rahbar' else youth.rahbar
        if not rahbar:
            return render(request, 'meetings/survey_form.html', {
                'youth': youth, 'survey': survey, 'questions': questions,
                'error': "Bu yoshga rahbar biriktirilmagan!"
            })

        meeting = Meeting.objects.create(
            rahbar=rahbar,
            youth=youth,
            date=timezone.now(),
            latitude=lat,
            longitude=lng,
            notes=notes_text,
            status='pending',
        )

        if photo_file:
            meeting.photo.save(f'meeting_{meeting.id}.jpg', photo_file, save=True)

        if youth.yetakchi:
            Verification.objects.create(meeting=meeting, verifier=youth.yetakchi)

        _notify_yetakchi_telegram(meeting, youth, rahbar)

        return redirect(f'/uchrashuvlar/{meeting.id}/natija/')

    return render(request, 'meetings/survey_form.html', {
        'youth': youth,
        'survey': survey,
        'questions': questions,
    })


# ===== Davlat Excel formati (12 ustun) =====
# 0:№  1:FISH  2:Tug'ilgan yili  3:Shahar/tuman  4:MFY nomi  5:Toifasi
# 6:Biriktirilgan tizm  7:ish joyi/lavozim  8:mas'ul FISH  9:mas'ul tel
# 10:Mahalla yoshlar yetakchisi  11:yetakchi Telegram raqami
IMPORT_HEADERS = [
    "№", "FISH", "Tug'ilgan yili, kuni", "Shahar tuman nomi", "MFY nomi", "Toifasi",
    "Biriktirilgan tizm", "ish joyi va lavozimi",
    "Biriktirilgan mas'ul FISH", "Biriktirilgan mas'ul tel raqami",
    "Biriktirilgan Mahalla yoshlar yetakchisi", "Mahalla yoshlar yetakchisining Telegram raqami",
]


def _parse_date(val):
    """Excel sana yoki matn -> date."""
    from datetime import datetime, date
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d.%m.%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _norm_category(val):
    """Toifa matnini kodga aylantiradi (kalit so'z bo'yicha)."""
    if not val:
        return 'other'
    s = str(val).strip().lower()
    if "yot g" in s or "yod g" in s:
        return 'yot_goya'
    if "sudlan" in s:
        return 'sudlangan'
    if "mehribonlik" in s:
        return 'mehribonlik'
    if "probatsiya" in s:
        return 'probatsiya'
    if "ijtimoiy" in s:
        return 'ijtimoiy'
    if "imkoniyat" in s:
        return 'imkoniyati_cheklangan'
    if "xavfli" in s or "xatarli" in s:
        return 'xatarli'
    if "yetim" in s or "yolg" in s:
        return 'yetim'
    if "ishsiz" in s:
        return 'ishsiz'
    return 'other'


def _split_name(full_name):
    parts = (full_name or '').strip().split()
    if not parts:
        return '', ''
    return parts[0], ' '.join(parts[1:])


_district_cache = {}
_org_cache = {}


def _get_district(name):
    name = (name or '').strip()
    if not name:
        return None
    if name in _district_cache:
        return _district_cache[name]
    d, _ = District.objects.get_or_create(name=name)
    _district_cache[name] = d
    return d


def _get_org(name, district):
    name = (name or '').strip()
    if not name:
        return None
    key = (name, district.id if district else None)
    if key in _org_cache:
        return _org_cache[key]
    o, _ = Organization.objects.get_or_create(
        name=name,
        district=district,
        defaults={'address': district.name if district else ''}
    )
    _org_cache[key] = o
    return o


def _get_or_create_staff(full_name, phone, role, district=None, org=None):
    """mas'ul -> rahbar, mahalla yetakchisi -> yetakchi. Telefon bilan bog'lanadi."""
    full_name = (full_name or '').strip()
    phone = (phone or '').strip()
    if not full_name and not phone:
        return None

    last9 = CustomUser.normalize_phone(phone)
    user = None
    if last9:
        user = CustomUser.objects.filter(phone_normalized__endswith=last9).exclude(phone_normalized='').first()
    if not user and full_name:
        first, last = _split_name(full_name)
        user = CustomUser.objects.filter(role=role, first_name__iexact=first, last_name__iexact=last).first()

    if user:
        # Bo'sh maydonlarni to'ldiramiz
        changed = False
        if phone and not user.phone:
            user.phone = phone; changed = True
        if district and not user.district_id:
            user.district = district; changed = True
        if org and not user.organization_id:
            user.organization = org; changed = True
        if changed:
            user.save()
        return user

    first, last = _split_name(full_name)
    if last9:
        base_username = f"{role}_{last9}"
    else:
        base_username = f"{role}_{first.lower() or 'user'}"
    username = base_username
    n = 1
    while CustomUser.objects.filter(username=username).exists():
        username = f"{base_username}_{n}"
        n += 1

    u = CustomUser(
        username=username, first_name=first, last_name=last,
        role=role, phone=phone, district=district, organization=org,
    )
    u.set_unusable_password()
    u.save()
    return u


@login_required
def import_excel(request):
    if not request.user.is_admin:
        return HttpResponse("Ruxsat yo'q", status=403)

    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(f)
        except Exception as e:
            return render(request, 'youth/import_result.html',
                          {'created': 0, 'errors': [f"Faylni o'qib bo'lmadi: {e}"]})
        ws = wb.active
        created = 0
        new_rahbar = 0
        new_yetakchi = 0
        errors = []
        _district_cache.clear()
        _org_cache.clear()

        rahbar_before = set(CustomUser.objects.filter(role='rahbar').values_list('id', flat=True))
        yetakchi_before = set(CustomUser.objects.filter(role='yetakchi').values_list('id', flat=True))

        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if not row:
                continue
            no = row[0]
            # Faqat raqamli № qatorlarini olamiz (sarlavha/bo'sh qatorlar tashlanadi)
            if no is None or not str(no).strip().isdigit():
                continue

            full_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not full_name:
                continue
            try:
                birth_date = _parse_date(row[2] if len(row) > 2 else None)
                district_name = str(row[3] or '').strip() if len(row) > 3 else ''
                mfy_name = str(row[4] or '').strip() if len(row) > 4 else ''
                category = _norm_category(row[5] if len(row) > 5 else None)
                tizm = str(row[6] or '').strip() if len(row) > 6 else ''
                lavozim = str(row[7] or '').strip() if len(row) > 7 else ''
                masul_name = str(row[8] or '').strip() if len(row) > 8 else ''
                masul_phone = str(row[9] or '').strip() if len(row) > 9 else ''
                yetakchi_name = str(row[10] or '').strip() if len(row) > 10 else ''
                yetakchi_phone = str(row[11] or '').strip() if len(row) > 11 else ''

                if not birth_date:
                    errors.append(f"{i}-qator ({full_name}): tug'ilgan sana noto'g'ri/yo'q")
                    continue

                district = _get_district(district_name)
                org = _get_org(mfy_name, district)
                rahbar = _get_or_create_staff(masul_name, masul_phone, 'rahbar', district, org)
                yetakchi = _get_or_create_staff(yetakchi_name, yetakchi_phone, 'yetakchi', district, org)

                if rahbar and rahbar.id not in rahbar_before:
                    new_rahbar += 1
                    rahbar_before.add(rahbar.id)
                if yetakchi and yetakchi.id not in yetakchi_before:
                    new_yetakchi += 1
                    yetakchi_before.add(yetakchi.id)

                address = ", ".join([x for x in [district_name, mfy_name] if x])
                notes_parts = []
                if tizm:
                    notes_parts.append(f"Biriktirilgan tizim: {tizm}")
                if lavozim:
                    notes_parts.append(f"Lavozim: {lavozim}")
                if row[5]:
                    notes_parts.append(f"Toifa: {str(row[5]).strip()}")

                Youth.objects.create(
                    full_name=full_name,
                    birth_date=birth_date,
                    address=address,
                    phone='',
                    category=category,
                    organization=org,
                    rahbar=rahbar,
                    yetakchi=yetakchi,
                    notes="\n".join(notes_parts),
                )
                created += 1
            except Exception as e:
                errors.append(f"{i}-qator ({full_name}): {e}")

        return render(request, 'youth/import_result.html', {
            'created': created,
            'new_rahbar': new_rahbar,
            'new_yetakchi': new_yetakchi,
            'errors': errors,
        })

    return render(request, 'youth/import.html')


@login_required
def sample_excel(request):
    """Davlat formatidagi namuna Excel (12 ustun)."""
    if not request.user.is_admin:
        return HttpResponse("Ruxsat yo'q", status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = " FISH"

    head_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    tel_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    head_font = Font(bold=True)
    for col, h in enumerate(IMPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = tel_fill if col == 12 else head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [5, 32, 14, 18, 16, 30, 18, 22, 28, 22, 28, 24]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    examples = [
        [1, "ABSATTAROV XUSNIDDIN SHUXRAT O'G'LI", "12.07.1998", "Bulung'ur tuman", "Gulzor MFY",
         "Yot g'oyalar ta'siriga tushib qolganlar", "Tuman(shahar) hokimi", "Bulung'ur tumani hokimi",
         "Astanov Samandar Mustafoqulovich", "94 374 74 72",
         "Valiyev Nabijon Aliyevich", "94 446 36 00"],
        [2, "BOYMURODOV TEMURBEK RAXMATULLA O'G'LI", "26.09.1995", "Bulung'ur tuman", "Nurli yo'l MFY",
         "Ilgari sudlanganlar (probatsiya va profilaktik hisobda)", "Tuman(shahar) hokimi", "Bulung'ur tumani hokimi",
         "Astanov Samandar Mustafoqulovich", "94 374 74 72",
         "Valiyev Nabijon Aliyevich", "94 446 36 00"],
        [3, "ABDULLAYEVA MAFTUNA UMID QIZI", "08.10.2006", "Bulung'ur tuman", "Guliston MFY",
         "Mehribonlik uyidan chiqqanlar", "Tuman(shahar) hokimi", "Bulung'ur tumani hokimi",
         "Astanov Samandar Mustafoqulovich", "94 374 74 72",
         "Valiyev Nabijon Aliyevich", "94 446 36 00"],
    ]
    for r, ex in enumerate(examples, start=2):
        for c, val in enumerate(ex, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 12:
                cell.fill = tel_fill

    # Yo'riqnoma varag'i
    ref = wb.create_sheet("Yo'riqnoma")
    lines = [
        ("YO'RIQNOMA — Excel to'ldirish qoidalari", True),
        ("", False),
        ("• 1-qatorda sarlavhalar, 2-qatordan boshlab ma'lumotlar yoziladi.", False),
        ("• MUHIM: 'Tuman' va 'MFY' nomi bo'yicha avtomatik yaratiladi — ID kerak emas.", False),
        ("• 'Biriktirilgan mas'ul' — RAHBAR sifatida, ism va telefon bilan avtomatik qo'shiladi.", False),
        ("• 'Mahalla yoshlar yetakchisi' — YETAKCHI sifatida qo'shiladi.", False),
        ("• SARIQ ustun (Telegram raqami) — eng muhimi! Yetakchi shu raqam bilan", False),
        ("  botga kirsa, tizim uni avtomatik taniydi.", False),
        ("• Sana formati: 12.07.1998 (kun.oy.yil).", False),
        ("", False),
        ("Toifa avtomatik aniqlanadi (matndan kalit so'z bo'yicha):", True),
    ]
    rr = 1
    for text, bold in lines:
        cell = ref.cell(row=rr, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, color="16A085")
        rr += 1
    for code, label in Youth.CATEGORY_CHOICES:
        ref.cell(row=rr, column=1, value=label)
        ref.cell(row=rr, column=2, value=f"-> {code}")
        rr += 1
    ref.column_dimensions['A'].width = 60
    ref.column_dimensions['B'].width = 20

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename=yoshlar_namuna.xlsx'
    wb.save(resp)
    return resp
